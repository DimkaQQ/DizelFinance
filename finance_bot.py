# -*- coding: utf-8 -*-
"""
DizelFinance Bot — Telegram бот для учёта финансов
Автор: Dimash | 2026
Обновлённая структура таблицы Planergo
"""
import os
import re
import logging
import base64
import json
import uuid
import threading
import time as _time_module
import requests
import xml.etree.ElementTree as ET
from datetime import datetime
from dotenv import load_dotenv
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from aiogram import Bot, Dispatcher, executor, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from flask import Flask, request, jsonify
from io import BytesIO

# ============================================================
# Сторонние библиотеки для парсинга
# ============================================================
try:
    import openpyxl
except ImportError:
    openpyxl = None
    logging.warning("openpyxl не установлен. Парсинг XLSX будет недоступен.")

try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None
    logging.warning("fitz (PyMuPDF) не установлен. Парсинг PDF будет недоступен.")

# ============================================================
# Глобальные хранилища (сессии — в памяти)
# ============================================================
pending_transactions = {}
pdf_sessions = {}

# ============================================================
# SQLite — постоянное хранилище черновиков
# ============================================================
import sqlite3
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "drafts.db")

def db_init():
    con = sqlite3.connect(DB_PATH)
    con.execute(
        "CREATE TABLE IF NOT EXISTS drafts ("
        "id TEXT PRIMARY KEY, user_id INTEGER, amount REAL, currency TEXT, "
        "rate REAL, amount_rub REAL, date TEXT, merchant TEXT, "
        "tx_type TEXT, created_at TEXT)"
    )
    con.commit()
    con.close()

db_init()

def drafts_get(user_id: int) -> list:
    con = sqlite3.connect(DB_PATH)
    rows = con.execute(
        "SELECT id,amount,currency,rate,amount_rub,date,merchant,tx_type "
        "FROM drafts WHERE user_id=? ORDER BY created_at DESC",
        (user_id,)
    ).fetchall()
    con.close()
    return [{"id":r[0],"a":r[1],"cur":r[2],"rate":r[3],"a_rub":r[4],
             "d":r[5],"m":r[6],"tx_type":r[7]} for r in rows]

def drafts_add(user_id: int, draft: dict):
    con = sqlite3.connect(DB_PATH)
    con.execute(
        "INSERT OR REPLACE INTO drafts VALUES (?,?,?,?,?,?,?,?,?)",
        (draft["id"], user_id, draft["a"], draft["cur"], draft["rate"],
         draft["a_rub"], draft["d"], draft["m"],
         draft.get("tx_type","Расход"), datetime.now().isoformat())
    )
    con.commit()
    con.close()

def drafts_remove(draft_id: str):
    con = sqlite3.connect(DB_PATH)
    con.execute("DELETE FROM drafts WHERE id=?", (draft_id,))
    con.commit()
    con.close()

def drafts_clear(user_id: int):
    con = sqlite3.connect(DB_PATH)
    con.execute("DELETE FROM drafts WHERE user_id=?", (user_id,))
    con.commit()
    con.close()

# ============================================================
# Настройка
# ============================================================
load_dotenv()
TELEGRAM_TOKEN   = os.getenv("TELEGRAM_TOKEN")
SHEET_URL        = os.getenv("SHEET_URL")
ADMIN_ID         = os.getenv("ADMIN_TELEGRAM_ID")
ALLOWED_IDS      = set(int(x.strip()) for x in os.getenv("ALLOWED_USER_IDS", "").split(",") if x.strip())
GEMINI_API_KEY   = os.getenv("GEMINI_API_KEY")
GEMINI_MODEL     = os.getenv("GEMINI_MODEL", "gemini-2.5-flash")
CLOUDFLARE_PROXY = os.getenv("CLOUDFLARE_PROXY", "https://gemini-proxy.dimash210775.workers.dev")

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

# ============================================================
# Google Sheets
# ============================================================
scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
creds = ServiceAccountCredentials.from_json_keyfile_name("finance-key.json", scope)
gc = gspread.authorize(creds)
sh = gc.open_by_url(SHEET_URL)

# ============================================================
# Курс валют — кеш 1 час, три источника
# ============================================================
_rate_cache: dict = {}
_RATE_TTL = 3600
_FALLBACK_RATES = {
    "USD": 90.0, "EUR": 98.0, "KZT": 0.19,
    "IDR": 0.0055, "VND": 0.0036,
}

def get_cbr_rate(currency: str) -> float:
    if currency == "RUB":
        return 1.0
    cached = _rate_cache.get(currency)
    if cached:
        rate, ts = cached
        if _time_module.time() - ts < _RATE_TTL:
            return rate
    rate = _fetch_rate(currency)
    _rate_cache[currency] = (rate, _time_module.time())
    return rate

def _fetch_rate(currency: str) -> float:
    try:
        resp = requests.get("https://api.exchangerate-api.com/v4/latest/RUB", timeout=4)
        if resp.status_code == 200:
            r = resp.json()["rates"].get(currency)
            if r and r > 0:
                return round(1.0 / r, 6)
    except Exception as e:
        logging.warning(f"exchangerate-api: {e}")

    try:
        resp = requests.get("https://www.cbr.ru/scripts/XML_daily.asp", timeout=4)
        root = ET.fromstring(resp.content)
        for valute in root.findall("Valute"):
            if valute.find("CharCode").text == currency:
                value   = valute.find("Value").text.replace(",", ".")
                nominal = int(valute.find("Nominal").text)
                return float(value) / nominal
    except Exception as e:
        logging.warning(f"cbr.ru: {e}")

    try:
        resp = requests.get("https://open.er-api.com/v6/latest/RUB", timeout=4)
        if resp.status_code == 200:
            r = resp.json().get("rates", {}).get(currency)
            if r and r > 0:
                return round(1.0 / r, 6)
    except Exception as e:
        logging.warning(f"open.er-api: {e}")

    fallback = _FALLBACK_RATES.get(currency, 1.0)
    logging.error(f"Все источники недоступны для {currency}, резервный курс: {fallback}")
    return fallback

# ============================================================
# Надёжный парсинг JSON из ответа Gemini
# ============================================================
def extract_json(text: str):
    text = text.strip()
    text = re.sub(r'```json\s*', '', text)
    text = re.sub(r'```\s*', '', text)
    text = text.strip()

    def find_matching_end(s, start, open_ch, close_ch):
        depth = 0
        in_string = False
        escape = False
        for i in range(start, len(s)):
            c = s[i]
            if escape:
                escape = False
                continue
            if c == '\\' and in_string:
                escape = True
                continue
            if c == '"':
                in_string = not in_string
                continue
            if in_string:
                continue
            if c == open_ch:
                depth += 1
            elif c == close_ch:
                depth -= 1
                if depth == 0:
                    return i
        return -1

    bracket_pos = text.find('[')
    brace_pos   = text.find('{')

    if bracket_pos != -1 and (brace_pos == -1 or bracket_pos < brace_pos):
        end = find_matching_end(text, bracket_pos, '[', ']')
        if end != -1:
            try:
                return json.loads(text[bracket_pos:end + 1])
            except json.JSONDecodeError as e:
                logging.warning(f"extract_json: ошибка парсинга массива: {e}")

    if brace_pos != -1:
        end = find_matching_end(text, brace_pos, '{', '}')
        if end != -1:
            try:
                return json.loads(text[brace_pos:end + 1])
            except json.JSONDecodeError as e:
                logging.warning(f"extract_json: ошибка парсинга объекта: {e}")

    return json.loads(text)

# ============================================================
# Claude API — замена Gemini
# ============================================================
def ask_gemini(prompt: str, image_bytes: bytes = None,
               mime_type: str = "image/jpeg", no_cache: bool = False) -> str:
    import hashlib

    ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
    CLAUDE_MODEL      = os.getenv("CLAUDE_MODEL", "claude-haiku-4-5-20251001")

    cache_key  = hashlib.md5(prompt.encode()).hexdigest() if (not image_bytes and not no_cache) else None
    cache_file = f"/tmp/claude_cache_{cache_key}.json" if cache_key else None

    if cache_file and os.path.exists(cache_file):
        try:
            with open(cache_file, encoding="utf-8") as f:
                cached_val = json.load(f)
                logging.info("Claude cache HIT")
                return cached_val
        except Exception:
            pass

    content = []
    if image_bytes:
        content.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": mime_type,
                "data": base64.b64encode(image_bytes).decode("utf-8")
            }
        })
    content.append({"type": "text", "text": prompt})

    payload = {
        "model": CLAUDE_MODEL,
        "max_tokens": 8000,
        "messages": [{"role": "user", "content": content}]
    }

    headers = {
        "x-api-key": ANTHROPIC_API_KEY,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json"
    }

    for attempt in range(5):
        try:
            resp = requests.post(
                "https://gemini-proxy.dimash210775.workers.dev/claude/v1/messages",
                json=payload, headers=headers, timeout=90
            )
            if resp.status_code == 200:
                data = resp.json()
                text = data["content"][0]["text"].strip()
                if cache_file:
                    with open(cache_file, "w", encoding="utf-8") as f:
                        json.dump(text, f, ensure_ascii=False)
                return text
            else:
                logging.warning(f"Claude error {resp.status_code}: {resp.text[:300]}")
                _time_module.sleep(5 * (attempt + 1))
        except Exception as e:
            logging.error(f"Claude exception (attempt {attempt + 1}): {e}")
            _time_module.sleep(5 * (attempt + 1))

    raise ValueError("Не удалось получить ответ от Claude после 5 попыток")

# ============================================================
# МАППИНГ СТАТЕЙ → ТАБЛИЦА
# ============================================================

# Ключевые слова для автоопределения категории "Джулиан" (питомцы/ветеринар/зоо)
JULIAN_KEYWORDS = [
    "ветеринар", "ветклиника", "вет ", "зоомагазин", "зоотовары", "зоо ",
    "petshop", "pet shop", "зоо-", "корм для", "ошейник", "поводок",
    "лоток", "наполнитель", "аквариум", "клетка для", "veterinary",
    "animal", "petstore", "зооуголок", "зоосалон", "груминг",
]

def is_julian_merchant(merchant: str) -> bool:
    """Проверяет, относится ли merchant к ветеринарным/зоо тратам."""
    m = merchant.lower()
    return any(kw in m for kw in JULIAN_KEYWORDS)

INCOME_ARTICLES = {
    "Зарплата":             "Поступления",
    "Прочие поступления":   "Поступления",
    "Премия и бонусы":      "Поступления",
    "Инвестиционный доход": "Поступления",
    # Движение активов
    "Портфель Екатерины":       "Движение активов",
    "Портфель Влада":           "Движение активов",
    "Портфель Ланы (подушка)":  "Движение активов",
    "Пенсионный план":          "Движение активов",
    "Инвестиции в бизнес":      "Движение активов",
}

EXPENSE_ARTICLES = {
    # Крупные расходы
    "Дети":         "Крупные расходы",
    "Джулиан":      "Крупные расходы",
    "Лана":         "Крупные расходы",
    "Образование":  "Крупные расходы",
    "Здоровье":     "Крупные расходы",
    "Путешествия":  "Крупные расходы",
    "Гаджеты":      "Крупные расходы",
    "Подарки":      "Крупные расходы",
    # Расходы
    "Продукты":     "Расходы",
    "Аренда":       "Расходы",
    "Транспорт":    "Расходы",
    "Развлечения":  "Расходы",
    "Здоровье ":    "Расходы",   # с пробелом — отдельная статья в таблице Расходы
    "Путешествия ": "Расходы",   # с пробелом — отдельная статья в таблице Расходы
    "Прочее":       "Расходы",
}

# Нормализованные имена (без trailing-пробела) для отображения
EXPENSE_ARTICLES_DISPLAY = {
    "Дети":         "Крупные расходы",
    "Джулиан":      "Крупные расходы",
    "Лана":         "Крупные расходы",
    "Образование":  "Крупные расходы",
    "Здоровье":     "Крупные расходы",
    "Путешествия":  "Крупные расходы",
    "Гаджеты":      "Крупные расходы",
    "Подарки":      "Крупные расходы",
    "Продукты":     "Расходы",
    "Аренда":       "Расходы",
    "Транспорт":    "Расходы",
    "Развлечения":  "Расходы",
    "Здоровье":     "Расходы",
    "Путешествия":  "Расходы",
    "Прочее":       "Расходы",
}

INCOME_BY_TABLE = {
    "Поступления": ["Зарплата", "Прочие поступления", "Премия и бонусы", "Инвестиционный доход"],
    "Движение активов":  ["Портфель Екатерины", "Портфель Влада", "Портфель Ланы (подушка)", "Пенсионный план", "Инвестиции в бизнес"],
}

EXPENSE_BY_TABLE = {
    "Крупные расходы": ["Дети", "Джулиан", "Лана", "Образование", "Здоровье", "Путешествия", "Гаджеты", "Подарки"],
    "Расходы":         ["Продукты", "Аренда", "Транспорт", "Развлечения", "Здоровье", "Путешествия", "Прочее"],
}

ALL_INCOME_ARTICLES  = list(INCOME_BY_TABLE["Поступления"]) + list(INCOME_BY_TABLE["Движение активов"])
ALL_EXPENSE_ARTICLES = list(EXPENSE_BY_TABLE["Крупные расходы"]) + list(EXPENSE_BY_TABLE["Расходы"])

MONTH_SHEETS = {
    1: "ЯНВАРЬ", 2: "ФЕВРАЛЬ",  3: "МАРТ",     4: "АПРЕЛЬ",
    5: "МАЙ",    6: "ИЮНЬ",     7: "ИЮЛЬ",     8: "АВГУСТ",
    9: "СЕНТЯБРЬ", 10: "ОКТЯБРЬ", 11: "НОЯБРЬ", 12: "ДЕКАБРЬ",
}

CURRENCIES       = ["RUB", "USD", "EUR", "KZT", "IDR", "VND"]
CURRENCY_SYMBOLS = {"RUB": "₽", "USD": "$", "EUR": "€", "KZT": "₸", "IDR": "Rp", "VND": "₫"}

# ============================================================
# Структура колонок таблицы (Planergo)
# ============================================================
TABLE_COLUMNS = {
    "Поступления":     (2,  12),
    "Крупные расходы": (15, 25),
    "Расходы":         (28, 38),
    "Движение активов":      (54, 64),
}

DATA_ROW_START = 28
DATA_ROW_END   = 39

TX_INCOME_DATE_COL     = 2
TX_INCOME_ARTICLE_COL  = 5
TX_INCOME_AMOUNT_COL   = 16
TX_EXPENSE_DATE_COL    = 35
TX_EXPENSE_ARTICLE_COL = 38
TX_EXPENSE_AMOUNT_COL  = 49

# ============================================================
# Логика листов месяцев
# ============================================================
def get_month_sheet_name(date_str: str) -> str:
    for fmt in ("%d.%m.%Y, %H:%M", "%d.%m.%Y,%H:%M", "%d.%m.%Y"):
        try:
            dt = datetime.strptime(date_str.strip(), fmt)
            return MONTH_SHEETS[dt.month]
        except ValueError:
            continue
    return MONTH_SHEETS[datetime.now().month]

def write_to_month_sheet(date_str: str, article: str, amount_rub: float, table_name: str):
    sheet_name = get_month_sheet_name(date_str)
    try:
        ws = sh.worksheet(sheet_name)
    except gspread.exceptions.WorksheetNotFound:
        logging.warning(f"Лист {sheet_name} не найден — пропускаем запись в месяц")
        return False

    col_name, col_fact = TABLE_COLUMNS.get(table_name, (None, None))
    if col_name is None:
        logging.warning(f"Неизвестная таблица: {table_name}")
        return False

    try:
        name_col_values = ws.col_values(col_name)
    except Exception as e:
        logging.error(f"Ошибка чтения столбца {col_name} листа {sheet_name}: {e}")
        return False

    target_row = None
    for row_idx in range(DATA_ROW_START, DATA_ROW_END + 1):
        cell_val = name_col_values[row_idx - 1] if row_idx <= len(name_col_values) else ""
        if cell_val.strip() == article.strip():
            target_row = row_idx
            break

    if target_row is None:
        logging.warning(f"Статья '{article}' не найдена в таблице '{table_name}' листа {sheet_name}")
        for row_idx in range(DATA_ROW_START, DATA_ROW_END + 1):
            cell_val = name_col_values[row_idx - 1] if row_idx <= len(name_col_values) else ""
            if not cell_val.strip():
                target_row = row_idx
                ws.update_cell(target_row, col_name, article)
                break

    if target_row is None:
        logging.error(f"Нет свободных строк в таблице '{table_name}' листа {sheet_name}")
        return False

    try:
        current_val = ws.cell(target_row, col_fact).value or "0"
        current_val = str(current_val).replace(" ", "").replace(",", ".").replace("₽", "").strip()
        current_amount = float(current_val) if current_val else 0.0
    except Exception:
        current_amount = 0.0

    new_amount = current_amount + amount_rub
    try:
        ws.update_cell(target_row, col_fact, round(new_amount, 2))
        logging.info(f"✅ Записано в {sheet_name}/{table_name}/{article}: {current_amount} + {amount_rub} = {new_amount}")
        return True
    except Exception as e:
        logging.error(f"Ошибка записи в ячейку {sheet_name} R{target_row}C{col_fact}: {e}")
        return False

def write_transaction_row(date_str: str, article: str, amount_rub: float,
                          currency: str, table_name: str, comment: str):
    sheet_name = get_month_sheet_name(date_str)
    try:
        ws = sh.worksheet(sheet_name)
    except Exception:
        return

    is_income = table_name in ("Поступления", "Движение активов")
    if is_income:
        date_col   = TX_INCOME_DATE_COL
        art_col    = TX_INCOME_ARTICLE_COL
        amount_col = TX_INCOME_AMOUNT_COL
    else:
        date_col   = TX_EXPENSE_DATE_COL
        art_col    = TX_EXPENSE_ARTICLE_COL
        amount_col = TX_EXPENSE_AMOUNT_COL

    try:
        col_vals = ws.col_values(date_col)
        next_row = 46
        for i, v in enumerate(col_vals[45:], start=46):
            if not str(v).strip():
                next_row = i
                break
        else:
            next_row = max(46, len(col_vals) + 1)

        date_short = date_str.split(",")[0].strip()
        ws.update_cell(next_row, date_col, date_short)
        ws.update_cell(next_row, art_col, article)
        ws.update_cell(next_row, amount_col, round(amount_rub, 2))
        logging.info(f"📝 Строка транзакции → {sheet_name} R{next_row} {article} {amount_rub}")
    except Exception as e:
        logging.error(f"Ошибка записи строки транзакции: {e}")

# ============================================================
# История и Угадывание статей
# ============================================================
_history_cache: dict = {"text": "", "ts": 0}
_HISTORY_TTL = 300

def _get_history_text() -> str:
    if _time_module.time() - _history_cache["ts"] < _HISTORY_TTL:
        return _history_cache["text"]
    try:
        ws = sh.worksheet("Транзакции")
        records = ws.get_all_records()
        last = records[-50:] if len(records) > 50 else records
        lines = [
            f"{str(r.get('Место','')).strip()} → {str(r.get('Статья','')).strip()}"
            for r in last
            if str(r.get("Место","")).strip() and str(r.get("Статья","")).strip()
        ]
        text = "\n".join(lines[-30:])
        _history_cache["text"] = text
        _history_cache["ts"]   = _time_module.time()
        return text
    except Exception as e:
        logging.warning(f"Не удалось загрузить историю: {e}")
        return _history_cache["text"]

def _resolve_article(article: str, tx_type: str) -> tuple:
    """
    Возвращает (article, table_name).
    Для расходов: сначала Крупные расходы, потом Расходы.
    """
    if tx_type == "Доход":
        for table, arts in INCOME_BY_TABLE.items():
            if article in arts:
                return article, table
        return "Прочие поступления", "Поступления"
    else:
        for table, arts in EXPENSE_BY_TABLE.items():
            if article in arts:
                return article, table
        return "Прочее", "Расходы"

def guess_article(merchant: str, amount: float, tx_type: str = "Расход", hint: str = "") -> tuple:
    # Ветеринар/зоо → всегда Джулиан
    if tx_type == "Расход" and is_julian_merchant(merchant):
        return "Джулиан", "Крупные расходы"

    history_text = _get_history_text()
    articles_str = json.dumps(
        ALL_INCOME_ARTICLES if tx_type == "Доход" else ALL_EXPENSE_ARTICLES,
        ensure_ascii=False
    )
    hint_line    = f"\nПодсказка банка: «{hint}»" if hint else ""
    history_line = ("История:\n" + history_text) if history_text else ""

    prompt = (
        f"Определи статью для транзакции.\n"
        f"Место: {merchant} | Сумма: {amount} | Тип: {tx_type}{hint_line}\n"
        f"{history_line}\n"
        f"Статьи: {articles_str}\n"
        f'Ответь ТОЛЬКО JSON: {{"article": "статья"}}\n'
        f"Выбирай только из списка. Если место уже в истории — используй ту же статью.\n"
        f"Важно: всё что связано с ветеринаром, зоомагазином, питомцами — статья 'Джулиан'."
    )
    try:
        result = ask_gemini(prompt)
        data   = extract_json(result)
        if isinstance(data, dict):
            return _resolve_article(data.get("article", ""), tx_type)
        raise ValueError("не dict")
    except Exception as e:
        logging.error(f"Ошибка угадывания статьи: {e}")
        if tx_type == "Доход":
            return "Прочие поступления", "Поступления"
        return "Прочее", "Расходы"

def guess_articles_batch(transactions: list) -> list:
    if not transactions:
        return []

    # Предобработка: зоо/вет → сразу Джулиан без AI
    pre_results = {}
    ai_indices  = []
    for i, tx in enumerate(transactions):
        if tx.get("tx_type", "Расход") == "Расход" and is_julian_merchant(tx.get("merchant", "")):
            pre_results[i] = ("Джулиан", "Крупные расходы")
        else:
            ai_indices.append(i)

    if not ai_indices:
        return [pre_results[i] for i in range(len(transactions))]

    history_text = _get_history_text()
    items = []
    for i in ai_indices:
        tx       = transactions[i]
        hint     = tx.get("category_hint", "")
        hint_str = f", подсказка банка: {hint}" if hint else ""
        items.append(
            f'{i}: merchant="{tx.get("merchant","")}", '
            f'amount={tx.get("amount",0)}, '
            f'type="{tx.get("tx_type","Расход")}"{hint_str}'
        )
    items_str    = "\n".join(items)
    expense_str  = json.dumps(ALL_EXPENSE_ARTICLES, ensure_ascii=False)
    income_str   = json.dumps(ALL_INCOME_ARTICLES, ensure_ascii=False)
    history_line = ("История:\n" + history_text) if history_text else ""

    prompt = (
        f"Определи статью для каждой транзакции.\n"
        f"Транзакции:\n{items_str}\n"
        f"Статьи расходов: {expense_str}\n"
        f"Статьи доходов: {income_str}\n"
        f"{history_line}\n"
        f"Ответь ТОЛЬКО JSON массивом, индексы совпадают с транзакциями:\n"
        f'[{{"index": 0, "article": "статья"}}, ...]\n'
        f"Правила:\n"
        f"- Выбирай ТОЛЬКО из предложенных статей\n"
        f'- Для type="Расход" — из статей расходов, для type="Доход" — из статей доходов\n'
        f"- Если место есть в истории — используй ту же статью\n"
        f"- Ветеринар, зоомагазин, питомцы → статья 'Джулиан'"
    )
    try:
        result   = ask_gemini(prompt)
        raw_list = extract_json(result)
        if not isinstance(raw_list, list):
            raise ValueError("не список")
        ai_results = {}
        for item in raw_list:
            idx     = item.get("index", -1)
            article = item.get("article", "")
            if idx in ai_indices:
                tx_type      = transactions[idx].get("tx_type", "Расход")
                ai_results[idx] = _resolve_article(article, tx_type)
        # Дефолт для не найденных
        for i in ai_indices:
            if i not in ai_results:
                tx_type = transactions[i].get("tx_type", "Расход")
                ai_results[i] = ("Прочие поступления", "Поступления") if tx_type == "Доход" else ("Прочее", "Расходы")
        # Собираем финальный список
        return [pre_results.get(i) or ai_results.get(i, ("Прочее", "Расходы")) for i in range(len(transactions))]
    except Exception as e:
        logging.error(f"Ошибка батч-угадывания: {e}")
        results = []
        for i, tx in enumerate(transactions):
            if i in pre_results:
                results.append(pre_results[i])
            else:
                results.append(guess_article(tx.get("merchant",""), tx.get("amount",0),
                                             tx.get("tx_type","Расход"), tx.get("category_hint","")))
        return results

# ============================================================
# Парсинг файлов (PDF, XLSX, SMS, Скриншот)
# ============================================================
def parse_pdf_transactions(pdf_base64: str) -> list:
    if not fitz:
        logging.error("fitz не установлен")
        return []
    try:
        pdf_bytes = base64.b64decode(pdf_base64)
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        all_transactions = []
        for i, page in enumerate(doc):
            pix       = page.get_pixmap(matrix=fitz.Matrix(150 / 72, 150 / 72))
            img_bytes = pix.tobytes("png")
            prompt = f"""Страница {i + 1} банковской выписки.
Извлеки ВСЕ транзакции. Для каждой верни:
- date: ДД.ММ.ГГГГ
- amount: число (положительное)
- currency: RUB/USD/EUR/KZT
- merchant: название магазина/места
- category_hint: категория трат если видна на странице, иначе ""
- tx_type: "Расход" или "Доход"
Игнорируй: пополнения счёта, переводы между своими счетами.
Ответ ТОЛЬКО JSON массивом:
[{{"date": "01.01.2024", "amount": 1500, "currency": "RUB", "merchant": "Пятёрочка", "tx_type": "Расход", "category_hint": "еда"}}]
Если транзакций нет — []."""
            result  = ask_gemini(prompt, image_bytes=img_bytes, mime_type="image/png")
            page_tx = extract_json(result)
            if isinstance(page_tx, list):
                all_transactions.extend(page_tx)
            elif isinstance(page_tx, dict):
                all_transactions.append(page_tx)
        doc.close()
        return all_transactions
    except Exception as e:
        logging.error(f"Ошибка парсинга PDF: {e}")
        return []

def parse_xlsx_transactions(file_bytes: bytes) -> list:
    if not openpyxl:
        logging.error("openpyxl не установлен")
        return []
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active

        header_row = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            row_str = [str(c).strip() if c else '' for c in row]
            if any('дата операции' in c.lower() for c in row_str if c):
                header_row = i
                break

        if header_row is None:
            logging.error("XLSX: не найдена строка заголовков")
            return []

        rows = list(ws.iter_rows(values_only=True))
        transactions = []

        for row in rows[header_row + 1:]:
            cells = [str(c).replace('\xa0', ' ').strip() if c else '' for c in row]
            if len(cells) < 13:
                continue
            date_val = cells[0]
            if not re.match(r'\d{2}\.\d{2}\.\d{4}', date_val):
                continue
            category = cells[4]
            desc     = cells[11]
            amount_str = cells[12].replace(' ', '').replace('\xa0', '').replace(',', '.')
            if not amount_str:
                continue
            try:
                amount_float = float(amount_str)
            except ValueError:
                continue

            place_match = re.search(
                r'место совершения операции:[^,]+/([^,\n]+?)(?:,|\s+MCC|$)',
                desc, re.IGNORECASE
            )
            if place_match:
                merchant = place_match.group(1).strip()
            else:
                merchant = desc[:50].strip()

            tx_type = 'Доход' if amount_float > 0 else 'Расход'

            transactions.append({
                'date':          date_val,
                'amount':        abs(amount_float),
                'currency':      'RUB',
                'merchant':      merchant,
                'tx_type':       tx_type,
                'category_hint': category,
            })

        logging.info(f"XLSX: найдено {len(transactions)} транзакций (без AI)")
        return transactions

    except Exception as e:
        logging.error(f"Ошибка парсинга XLSX: {e}")
        return []

def parse_screenshot_transactions(image_bytes: bytes, mime_type: str = "image/jpeg") -> list:
    prompt = """Это скриншот банковского приложения. Извлеки ВСЕ транзакции.
Для каждой верни:
- date: ДД.ММ.ГГГГ (если нет года — добавь текущий)
- amount: число (положительное)
- currency: RUB/USD/EUR/KZT/IDR/VND (определи по символу ₽/$/€/₸/Rp/₫)
- merchant: название места или описание
- tx_type: "Расход" (списание) или "Доход" (зачисление)
- category_hint: если на скриншоте РЯДОМ с транзакцией написана категория — напиши её, иначе ""
Ответ ТОЛЬКО JSON массивом:
[{"date": "01.01.2024", "amount": 1500, "currency": "RUB", "merchant": "Пятёрочка", "tx_type": "Расход", "category_hint": "Еда"}]
Если нет транзакций — [].
Игнорируй: баланс, заголовки, рекламу."""
    try:
        result = ask_gemini(prompt, image_bytes=image_bytes, mime_type=mime_type)
        transactions = extract_json(result)
        if isinstance(transactions, list):
            return transactions
        if isinstance(transactions, dict):
            return [transactions]
        return []
    except Exception as e:
        logging.error(f"Ошибка парсинга скриншота: {e}")
        return []

def parse_sms_transaction(sms_text: str) -> dict | None:
    prompt = f"""Извлеки данные транзакции из банковского SMS.
SMS: {sms_text}
Ответь ТОЛЬКО в формате JSON без markdown:
{{"amount": 0.0, "currency": "RUB", "merchant": "название места или описание", "tx_type": "Расход или Доход", "date": "ДД.ММ.ГГГГ или пустая строка"}}
Правила:
- tx_type = "Доход" если это пополнение/зачисление/перевод ПОЛУЧЕН
- tx_type = "Расход" если это списание/покупка/оплата
- amount всегда положительное число
- если дата не указана — пустая строка
Если это НЕ банковское SMS с транзакцией — верни {{"error": "not_transaction"}}"""
    try:
        result = ask_gemini(prompt, no_cache=True)
        data   = extract_json(result)
        if isinstance(data, dict) and data.get("error") == "not_transaction":
            return None
        return data if isinstance(data, dict) else None
    except Exception as e:
        logging.error(f"Ошибка парсинга SMS: {e}")
        return None

def get_existing_transactions() -> set:
    try:
        ws      = sh.worksheet("Транзакции")
        records = ws.get_all_records()
        existing = set()
        for rec in records:
            date   = str(rec.get("Дата", "")).split(",")[0].strip()
            amount = str(rec.get("Сумма", "")).strip()
            # Место нет в таблице — матчим только по дата+сумма
            existing.add(f"{date}|{amount}")
        return existing
    except Exception as e:
        logging.error(f"Ошибка получения транзакций: {e}")
        return set()

# ============================================================
# Инициализация бота
# ============================================================
bot     = Bot(token=TELEGRAM_TOKEN)
storage = MemoryStorage()
dp      = Dispatcher(bot, storage=storage)

# ============================================================
# Состояния FSM
# ============================================================
class TransactionForm(StatesGroup):
    waiting_for_action = State()
    tx_type            = State()
    table_choice       = State()
    article_choice     = State()
    amount             = State()
    currency           = State()
    date               = State()
    final_confirmation = State()
    edit_amount        = State()
    edit_currency      = State()

class PDFForm(StatesGroup):
    reviewing = State()

# ============================================================
# Клавиатуры
# ============================================================
def main_menu_kb():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.row("➕ Новая транзакция")
    kb.row("📋 Мои транзакции", "📊 Статистика")
    kb.row("📥 Отложенные", "⚙️ Настройки")
    return kb

def tx_type_kb():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    kb.add("💸 Расход", "💰 Доход")
    kb.add("⏪ Назад")
    return kb

def table_choice_kb(tx_type: str):
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    if tx_type == "Доход":
        kb.add("📥 Поступления")
        kb.add("🏦 Движение активов")
    else:
        kb.add("💳 Крупные расходы")
        kb.add("🛒 Расходы")
    kb.add("⏪ Назад")
    return kb

TABLE_LABEL_MAP = {
    "📥 Поступления":  "Поступления",
    "🏦 Движение активов":   "Движение активов",
    "💳 Крупные расходы": "Крупные расходы",
    "🛒 Расходы":      "Расходы",
}

def articles_kb(articles: list):
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    for art in articles:
        kb.add(art)
    kb.add("⏪ Назад")
    return kb

def currencies_kb():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    kb.add(*CURRENCIES)
    kb.add("⏪ Назад")
    return kb

def back_kb():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    kb.add("⏪ Назад")
    return kb

def skip_kb():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    kb.add("Пропустить")
    kb.add("⏪ Назад")
    return kb

def confirmation_kb():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    kb.add("✅ Записать")
    kb.add("✏️ Изменить статью")
    kb.add("🔢 Изменить сумму/валюту")
    kb.add("❌ Отменить")
    return kb

def pdf_action_kb():
    kb = types.InlineKeyboardMarkup(row_width=2)
    kb.add(
        types.InlineKeyboardButton("✅ Записать все", callback_data="pdf|all"),
        types.InlineKeyboardButton("👀 Просмотреть",  callback_data="pdf|review"),
    )
    kb.add(types.InlineKeyboardButton("❌ Отменить", callback_data="pdf|cancel"))
    return kb

def pdf_item_kb(idx: int, total: int):
    kb = types.InlineKeyboardMarkup(row_width=2)
    kb.add(
        types.InlineKeyboardButton("✅ Записать",   callback_data=f"pdfi|save|{idx}"),
        types.InlineKeyboardButton("⏭ Пропустить", callback_data=f"pdfi|skip|{idx}"),
    )
    kb.add(
        types.InlineKeyboardButton("✏️ Статья",     callback_data=f"pdfi|edit_cat|{idx}"),
        types.InlineKeyboardButton("🔢 Сумма/Вал.", callback_data=f"pdfi|edit_amt|{idx}"),
    )
    if idx + 1 < total:
        kb.add(types.InlineKeyboardButton(f"→ Следующая ({idx + 2}/{total})", callback_data=f"pdfi|next|{idx}"))
    else:
        kb.add(types.InlineKeyboardButton("🏁 Завершить", callback_data="pdfi|done|0"))
    return kb

# ============================================================
# Превью транзакции
# ============================================================
def build_preview(data: dict) -> str:
    currency   = data.get("currency", "RUB")
    amount     = data.get("amount", 0)
    rate       = data.get("rate", 1.0)
    amount_rub = data.get("amount_rub", amount)
    symbol     = CURRENCY_SYMBOLS.get(currency, currency)
    tx_type    = data.get("tx_type", "Расход")
    tx_icon    = "💰" if tx_type == "Доход" else "💸"
    article    = data.get("article", "")
    table_name = data.get("table_name", "")

    preview = (
        f"📝 <b>Предварительный просмотр:</b>\n"
        f"{tx_icon} {tx_type}\n"
        f"📅 Дата: <code>{data.get('date', '')}</code>\n"
        f"📂 Таблица: <code>{table_name}</code>\n"
        f"📋 Статья: <code>{article}</code>\n"
    )
    if currency == "RUB":
        preview += f"💰 Сумма: <code>{float(amount):,.0f}</code> ₽\n"
    else:
        preview += (
            f"💰 Сумма: <code>{float(amount):,.2f}</code> {symbol}\n"
            f"💱 Курс ЦБ: <code>{float(rate):,.4f}</code> ₽/{symbol}\n"
            f"🔄 В рублях: <code>{float(amount_rub):,.0f}</code> ₽\n"
        )
    return preview

def build_pdf_tx_preview(tx: dict, idx: int, total: int) -> str:
    currency   = tx.get("currency", "RUB")
    amount     = tx.get("amount", 0)
    symbol     = CURRENCY_SYMBOLS.get(currency, currency)
    article    = tx.get("article", "Прочее")
    table_name = tx.get("table_name", "Расходы")
    tx_type    = tx.get("tx_type", "Расход")
    tx_icon    = "💰" if tx_type == "Доход" else "💸"
    hint       = tx.get("category_hint", "")

    text = (
        f"<b>#{idx + 1} из {total}</b>\n"
        f"{tx_icon} <b>{tx.get('merchant', '')}</b>\n"
        f"💰 {float(amount):,.2f} {symbol}\n"
        f"📅 {tx.get('date', '')}\n"
        f"📂 {table_name} → {article}"
    )
    if hint:
        text += f"\n💡 Подсказка банка: <i>{hint}</i>"
    if tx.get("is_duplicate"):
        text += "\n⚠️ <i>Возможно уже записана</i>"
    return text

# ============================================================
# Запись транзакции: в лист «Транзакции» + в лист месяца
# ============================================================
async def save_transaction_to_sheets(data: dict):
    article    = data.get("article", "")
    table_name = data.get("table_name", "")
    tx_type    = data.get("tx_type", "Расход")
    currency   = data.get("currency", "RUB")
    amount     = data.get("amount", 0)
    rate       = data.get("rate", 1.0)
    amount_rub = data.get("amount_rub", amount)
    date_str   = data.get("date", "")

    ws = sh.worksheet("Транзакции")
    try:
        headers = ws.row_values(1)
        if headers and len(headers) > 1 and headers[1] == "Категория":
            ws.update("A1:H1", [["Дата", "Таблица", "Статья", "Сумма", "Валюта",
                                  "Курс", "Сумма в Руб", "Тип"]])
            logging.info("Заголовки листа Транзакции обновлены")
    except Exception as e:
        logging.warning(f"Не удалось проверить/обновить заголовки: {e}")

    new_row = [
        date_str,
        table_name,
        article,
        amount,
        currency,
        rate if currency != "RUB" else "",
        amount_rub,
        tx_type,
    ]
    ws.append_row(new_row)

    write_to_month_sheet(date_str, article, float(amount_rub), table_name)
    write_transaction_row(date_str, article, float(amount_rub), currency,
                         table_name, data.get("comment", ""))

async def save_transactions_batch(transactions: list) -> int:
    from collections import defaultdict
    if not transactions:
        return 0

    ws_tx = sh.worksheet("Транзакции")

    try:
        headers = ws_tx.row_values(1)
        if not headers or len(headers) < 7 or (len(headers) > 1 and headers[1] == "Категория"):
            ws_tx.update("A1:H1", [["Дата", "Таблица", "Статья", "Сумма",
                                     "Валюта", "Курс", "Сумма в Руб", "Тип"]])
    except Exception as e:
        logging.warning(f"Заголовки: {e}")

    new_rows = []
    for data in transactions:
        currency = data.get("currency", "RUB")
        new_rows.append([
            data.get("date", ""),
            data.get("table_name", ""),
            data.get("article", ""),
            data.get("amount", 0),
            currency,
            data.get("rate", "") if currency != "RUB" else "",
            data.get("amount_rub", data.get("amount", 0)),
            data.get("tx_type", "Расход"),
        ])

    if new_rows:
        ws_tx.append_rows(new_rows, value_input_option="USER_ENTERED")
        logging.info(f"✅ Батч: {len(new_rows)} строк за 1 вызов")

    month_sums = defaultdict(float)
    for data in transactions:
        key = (
            get_month_sheet_name(data.get("date", "")),
            data.get("table_name", ""),
            data.get("article", ""),
        )
        month_sums[key] += float(data.get("amount_rub", data.get("amount", 0)))

    # Маппинг имени листа → дата для write_to_month_sheet
    MONTH_TO_DATE = {v: f"01.{k:02d}.{datetime.now().year}"
                     for k, v in MONTH_SHEETS.items()}

    ws_cache = {}
    for (sheet_name, table_name, article), total_rub in month_sums.items():
        if not table_name or not article:
            continue
        try:
            if sheet_name not in ws_cache:
                ws_cache[sheet_name] = sh.worksheet(sheet_name)
            fake_date = MONTH_TO_DATE.get(sheet_name, datetime.now().strftime("%d.%m.%Y"))
            write_to_month_sheet(fake_date, article, total_rub, table_name)
        except gspread.exceptions.WorksheetNotFound:
            logging.warning(f"Лист {sheet_name} не найден")
        except Exception as e:
            logging.error(f"Ошибка {sheet_name}/{article}: {e}")

    return len(new_rows)

# ============================================================
# Уведомление админу
# ============================================================
async def notify_admin(message_text: str, user: types.User = None):
    if not ADMIN_ID:
        return
    user_info = ""
    if user:
        username  = f"@{user.username}" if user.username else ""
        user_info = f"\n👤 {user.full_name} {username} (ID: {user.id})"
    try:
        await bot.send_message(ADMIN_ID, f"🔔 <b>Новая транзакция</b>\n{message_text}{user_info}", parse_mode="HTML")
    except Exception as e:
        logging.error(f"Ошибка уведомления админу: {e}")

# ============================================================
# /start
# ============================================================
@dp.message_handler(commands=["start"], state="*")
async def cmd_start(message: types.Message, state: FSMContext):
    if ALLOWED_IDS and message.from_user.id not in ALLOWED_IDS:
        await message.answer("🔒 У вас нет доступа к этому боту.")
        return
    await state.finish()
    await TransactionForm.waiting_for_action.set()
    await message.answer(
        "👋 <b>Добро пожаловать в DizelFinance!</b>\n"
        "📱 <b>Способы записи транзакций:</b>\n"
        "— Отправь <b>скриншот</b> банка → AI распознает\n"
        "— Отправь <b>PDF выписку</b> → автопарсинг\n"
        "— Отправь <b>Excel выписку</b> → автопарсинг\n"
        "— Вставь <b>текст SMS</b> → автораспознавание\n"
        "— «➕ Новая транзакция» → ручной ввод\n"
        "\nТранзакции записываются:\n"
        "✅ В лист «Транзакции»\n"
        "✅ В лист текущего месяца (ЯНВАРЬ, ФЕВРАЛЬ...)\n"
        "Выберите действие:",
        parse_mode="HTML", reply_markup=main_menu_kb()
    )

# ============================================================
# Обработка документов (PDF и XLSX)
# ============================================================
@dp.message_handler(content_types=types.ContentType.DOCUMENT, state="*")
async def handle_document(message: types.Message, state: FSMContext):
    if ALLOWED_IDS and message.from_user.id not in ALLOWED_IDS:
        return
    fname = (message.document.file_name or "").lower()
    if fname.endswith(('.xlsx', '.xls')):
        await handle_xlsx(message, state)
    elif fname.endswith('.pdf'):
        await handle_pdf(message, state)
    else:
        await message.answer("Пожалуйста отправьте PDF или Excel файл.")

async def handle_xlsx(message: types.Message, state: FSMContext):
    await message.answer("⏳ Скачиваю Excel выписку...")
    try:
        file       = await bot.get_file(message.document.file_id)
        file_bytes = await bot.download_file(file.file_path)
        xlsx_bytes = file_bytes.read()
        await message.answer("🤖 Анализирую выписку...")
        transactions = parse_xlsx_transactions(xlsx_bytes)
        if not transactions:
            await message.answer("❌ Не удалось найти транзакции в файле.", reply_markup=main_menu_kb())
            return
        existing = get_existing_transactions()
        await message.answer(f"📊 Найдено {len(transactions)} транзакций. Определяю статьи...")
        article_results = guess_articles_batch(transactions)
        enriched = _enrich_transactions(transactions, article_results, existing)
        _store_session(message.from_user.id, enriched)
        await _send_session_summary(message, enriched, "Excel")
    except Exception as e:
        logging.error(f"Ошибка обработки XLSX: {e}")
        await message.answer(f"❌ Ошибка: {e}", reply_markup=main_menu_kb())

async def handle_pdf(message: types.Message, state: FSMContext):
    await message.answer("⏳ Скачиваю и читаю выписку...")
    try:
        file       = await bot.get_file(message.document.file_id)
        file_bytes = await bot.download_file(file.file_path)
        pdf_base64 = base64.b64encode(file_bytes.read()).decode('utf-8')
        await message.answer("🤖 AI анализирует транзакции...")
        transactions = parse_pdf_transactions(pdf_base64)
        if not transactions:
            await message.answer("❌ Не удалось найти транзакции в файле.")
            return
        existing = get_existing_transactions()
        await message.answer(f"📊 Найдено {len(transactions)} транзакций. Определяю статьи через AI...")
        article_results = guess_articles_batch(transactions)
        enriched = _enrich_transactions(transactions, article_results, existing)
        _store_session(message.from_user.id, enriched)
        await _send_session_summary(message, enriched, "PDF")
    except Exception as e:
        logging.error(f"Ошибка обработки PDF: {e}")
        await message.answer(f"❌ Ошибка при обработке PDF: {e}")

def _enrich_transactions(transactions: list, article_results: list, existing: set) -> list:
    enriched = []
    for tx, (article, table_name) in zip(transactions, article_results):
        date_part  = str(tx.get("date", "")).split(",")[0].strip()
        amount_str = str(tx.get("amount", ""))
        # Было: f"{date_part}|{amount_str}|{merchant_key}"
        is_duplicate = f"{date_part}|{amount_str}" in existing
        currency     = tx.get("currency", "RUB")
        rate         = get_cbr_rate(currency)
        amount_rub   = round(float(tx.get("amount", 0)) * rate, 2)
        enriched.append({
            **tx,
            "article":       article,
            "table_name":    table_name,
            "rate":          rate,
            "amount_rub":    amount_rub,
            "is_duplicate":  is_duplicate,
            "category_hint": tx.get("category_hint", ""),
        })
    return enriched

def _store_session(user_id: int, enriched: list):
    pdf_sessions[user_id] = {
        "transactions":  enriched,
        "current_idx":   0,
        "saved_count":   0,
        "skipped_count": 0,
    }

async def _send_session_summary(message: types.Message, enriched: list, source_label: str):
    dup   = sum(1 for t in enriched if t["is_duplicate"])
    new_c = len(enriched) - dup
    await message.answer(
        f"✅ <b>{source_label} обработан!</b>\n"
        f"📄 Транзакций: {len(enriched)}\n"
        f"🆕 Новых: {new_c}\n"
        f"⚠️ Возможных дубликатов: {dup}\n"
        f"Что делаем?",
        parse_mode="HTML", reply_markup=pdf_action_kb()
    )

# ============================================================
# PDF/XLSX — действия
# ============================================================
@dp.callback_query_handler(lambda c: c.data.startswith("pdf|"), state="*")
async def pdf_action_handler(callback: types.CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    action  = callback.data.split("|")[1]
    session = pdf_sessions.get(user_id)
    if not session:
        await callback.message.edit_text("❌ Сессия устарела. Загрузите файл заново.")
        return
    if action == "cancel":
        pdf_sessions.pop(user_id, None)
        await callback.message.edit_text("❌ Отменено.")
        await callback.message.answer("Выберите действие:", reply_markup=main_menu_kb())
    elif action == "all":
        await callback.message.edit_text("⏳ Записываю все транзакции...")
        try:
            batch = [
                {
                    "date":       tx.get("date", ""),
                    "article":    tx.get("article", "Прочее"),
                    "table_name": tx.get("table_name", "Расходы"),
                    "amount":     tx.get("amount", 0),
                    "currency":   tx.get("currency", "RUB"),
                    "rate":       tx.get("rate", 1.0),
                    "amount_rub": tx.get("amount_rub", tx.get("amount", 0)),
                    "tx_type":    tx.get("tx_type", "Расход"),
                }
                for tx in session["transactions"]
            ]
            saved = await save_transactions_batch(batch)
            pdf_sessions.pop(user_id, None)
            await callback.message.answer(
                f"✅ Записано {saved} транзакций!", reply_markup=main_menu_kb()
            )
        except Exception as e:
            logging.error(f"Ошибка батч-записи: {e}")
            await callback.message.answer(f"❌ Ошибка: {e}", reply_markup=main_menu_kb())
    elif action == "review":
        session["current_idx"] = 0
        await PDFForm.reviewing.set()
        await callback.message.edit_text("👀 Просматриваем по одной...")
        await show_pdf_transaction(callback.message, user_id, 0)
        await callback.answer()

async def show_pdf_transaction(message: types.Message, user_id: int, idx: int):
    session = pdf_sessions.get(user_id)
    if not session:
        await message.answer("❌ Сессия завершена.", reply_markup=main_menu_kb())
        return
    transactions = session["transactions"]
    if idx >= len(transactions):
        saved   = session.get("saved_count", 0)
        skipped = session.get("skipped_count", 0)
        pdf_sessions.pop(user_id, None)
        await message.answer(
            f"🏁 <b>Готово!</b>\n✅ Записано: {saved}\n⏭ Пропущено: {skipped}",
            parse_mode="HTML", reply_markup=main_menu_kb()
        )
        return
    tx   = transactions[idx]
    text = build_pdf_tx_preview(tx, idx, len(transactions))
    await message.answer(text, parse_mode="HTML", reply_markup=pdf_item_kb(idx, len(transactions)))

@dp.callback_query_handler(lambda c: c.data.startswith("pdfi|"), state="*")
async def pdf_item_handler(callback: types.CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    parts   = callback.data.split("|")
    action  = parts[1]
    idx     = int(parts[2])
    session = pdf_sessions.get(user_id)
    if not session:
        await callback.message.edit_text("❌ Сессия устарела.")
        await callback.answer()
        return

    transactions = session["transactions"]

    async def _remove_kb():
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except Exception:
            pass

    if action == "done":
        saved   = session.get("saved_count", 0)
        skipped = session.get("skipped_count", 0)
        pdf_sessions.pop(user_id, None)
        await callback.message.edit_text(
            f"🏁 <b>Готово!</b>\n✅ Записано: {saved}\n⏭ Пропущено: {skipped}",
            parse_mode="HTML"
        )
        await callback.message.answer("Выберите действие:", reply_markup=main_menu_kb())
        await state.finish()
        await TransactionForm.waiting_for_action.set()
    elif action == "save":
        tx = transactions[idx]
        try:
            await save_transaction_to_sheets({
                "date":       tx.get("date", ""),
                "article":    tx.get("article", "Прочее"),
                "table_name": tx.get("table_name", "Расходы"),
                "amount":     tx.get("amount", 0),
                "currency":   tx.get("currency", "RUB"),
                "rate":       tx.get("rate", 1.0),
                "amount_rub": tx.get("amount_rub", tx.get("amount", 0)),
                "comment":    tx.get("merchant", ""),
                "tx_type":    tx.get("tx_type", "Расход"),
            })
            session["saved_count"] = session.get("saved_count", 0) + 1
            await callback.answer("✅ Записано!")
        except Exception as e:
            await callback.answer(f"❌ Ошибка: {e}")
        await _remove_kb()
        await show_pdf_transaction(callback.message, user_id, idx + 1)
    elif action == "skip":
        session["skipped_count"] = session.get("skipped_count", 0) + 1
        await callback.answer("⏭ Пропущено")
        await _remove_kb()
        await show_pdf_transaction(callback.message, user_id, idx + 1)
    elif action == "next":
        await _remove_kb()
        await show_pdf_transaction(callback.message, user_id, idx + 1)
    elif action == "edit_cat":
        tx      = transactions[idx]
        tx_type = tx.get("tx_type", "Расход")
        await state.update_data(
            amount=float(tx.get("amount", 0)),
            currency=tx.get("currency", "RUB"),
            rate=tx.get("rate", 1.0),
            amount_rub=tx.get("amount_rub", tx.get("amount", 0)),
            date=tx.get("date", ""),
            comment=tx.get("merchant", ""),
            tx_type=tx_type,
            from_pdf=True,
            pdf_idx=idx,
        )
        await _remove_kb()
        await TransactionForm.table_choice.set()
        await callback.message.answer(
            f"✏️ Редактируем статью для:\n<b>{tx.get('merchant','')}</b> — {tx.get('amount',0)} {tx.get('currency','RUB')}\nВыберите таблицу:",
            parse_mode="HTML",
            reply_markup=table_choice_kb(tx_type)
        )
    elif action == "edit_amt":
        tx = transactions[idx]
        await state.update_data(
            pdf_idx=idx,
            from_pdf=True,
            currency=tx.get("currency", "RUB"),
            date=tx.get("date", ""),
            comment=tx.get("merchant", ""),
            tx_type=tx.get("tx_type", "Расход"),
            article=tx.get("article", ""),
            table_name=tx.get("table_name", ""),
        )
        await _remove_kb()
        await TransactionForm.edit_amount.set()
        symbol = CURRENCY_SYMBOLS.get(tx.get("currency","RUB"), tx.get("currency","RUB"))
        await callback.message.answer(
            f"🔢 Текущая сумма: <b>{tx.get('amount',0)} {symbol}</b>\nВведите новую сумму:",
            parse_mode="HTML",
            reply_markup=back_kb()
        )
        await callback.answer()

# ============================================================
# Редактирование суммы из PDF/скриншота
# ============================================================
@dp.message_handler(state=TransactionForm.edit_amount)
async def process_edit_amount(message: types.Message, state: FSMContext):
    if message.text == "⏪ Назад":
        data    = await state.get_data()
        idx     = data.get("pdf_idx", 0)
        user_id = message.from_user.id
        await state.finish()
        await PDFForm.reviewing.set()
        await show_pdf_transaction(message, user_id, idx)
        return
    try:
        new_amount = float(message.text.replace(",", ".").replace(" ", ""))
        if new_amount <= 0:
            raise ValueError
    except ValueError:
        await message.answer("Введите корректную сумму:", reply_markup=back_kb())
        return
    await state.update_data(new_amount=new_amount)
    await TransactionForm.edit_currency.set()
    await message.answer("Выберите валюту:", reply_markup=currencies_kb())

@dp.message_handler(state=TransactionForm.edit_currency)
async def process_edit_currency(message: types.Message, state: FSMContext):
    if message.text == "⏪ Назад":
        await TransactionForm.edit_amount.set()
        await message.answer("Введите сумму:", reply_markup=back_kb())
        return
    if message.text not in CURRENCIES:
        await message.answer("Выберите валюту из списка:", reply_markup=currencies_kb())
        return
    data       = await state.get_data()
    idx        = data.get("pdf_idx", 0)
    new_amount = data.get("new_amount", 0)
    currency   = message.text
    rate       = get_cbr_rate(currency)
    amount_rub = round(new_amount * rate, 2)
    user_id    = message.from_user.id
    session    = pdf_sessions.get(user_id)
    if session and idx < len(session["transactions"]):
        session["transactions"][idx]["amount"]     = new_amount
        session["transactions"][idx]["currency"]   = currency
        session["transactions"][idx]["rate"]       = rate
        session["transactions"][idx]["amount_rub"] = amount_rub
    await state.finish()
    await PDFForm.reviewing.set()
    await message.answer("✅ Сумма обновлена!", reply_markup=types.ReplyKeyboardRemove())
    await show_pdf_transaction(message, user_id, idx)

# ============================================================
# Главное меню → новая транзакция
# ============================================================
@dp.message_handler(lambda m: m.text == "➕ Новая транзакция", state="*")
async def new_transaction(message: types.Message, state: FSMContext):
    await state.finish()
    await TransactionForm.tx_type.set()
    await message.answer("Тип операции:", reply_markup=tx_type_kb())

# ============================================================
# Тип операции
# ============================================================
@dp.message_handler(state=TransactionForm.tx_type)
async def process_tx_type(message: types.Message, state: FSMContext):
    if message.text == "⏪ Назад":
        await TransactionForm.waiting_for_action.set()
        await message.answer("Выберите действие:", reply_markup=main_menu_kb())
        return
    if message.text not in ["💸 Расход", "💰 Доход"]:
        await message.answer("Выберите тип:", reply_markup=tx_type_kb())
        return
    tx_type = "Расход" if message.text == "💸 Расход" else "Доход"
    await state.update_data(tx_type=tx_type)
    await TransactionForm.table_choice.set()
    await message.answer("Выберите таблицу:", reply_markup=table_choice_kb(tx_type))

# ============================================================
# Выбор таблицы
# ============================================================
@dp.message_handler(state=TransactionForm.table_choice)
async def process_table_choice(message: types.Message, state: FSMContext):
    data = await state.get_data()
    if message.text == "⏪ Назад":
        if data.get("from_pdf"):
            idx = data.get("pdf_idx", 0)
            await state.finish()
            await PDFForm.reviewing.set()
            await show_pdf_transaction(message, message.from_user.id, idx)
            return
        await TransactionForm.tx_type.set()
        await message.answer("Тип операции:", reply_markup=tx_type_kb())
        return
    table_name = TABLE_LABEL_MAP.get(message.text)
    if not table_name:
        await message.answer("Выберите таблицу:", reply_markup=table_choice_kb(data.get("tx_type", "Расход")))
        return
    await state.update_data(table_name=table_name)
    tx_type  = data.get("tx_type", "Расход")
    articles = INCOME_BY_TABLE.get(table_name, []) if tx_type == "Доход" else EXPENSE_BY_TABLE.get(table_name, [])
    await TransactionForm.article_choice.set()
    await message.answer(f"Выберите статью ({table_name}):", reply_markup=articles_kb(articles))

# ============================================================
# Выбор статьи
# ============================================================
@dp.message_handler(state=TransactionForm.article_choice)
async def process_article_choice(message: types.Message, state: FSMContext):
    data       = await state.get_data()
    tx_type    = data.get("tx_type", "Расход")
    table_name = data.get("table_name", "")
    if message.text == "⏪ Назад":
        if data.get("from_pdf"):
            idx = data.get("pdf_idx", 0)
            await state.finish()
            await PDFForm.reviewing.set()
            await show_pdf_transaction(message, message.from_user.id, idx)
            return
        await TransactionForm.table_choice.set()
        await message.answer("Выберите таблицу:", reply_markup=table_choice_kb(tx_type))
        return
    valid = INCOME_BY_TABLE.get(table_name, []) if tx_type == "Доход" else EXPENSE_BY_TABLE.get(table_name, [])
    if message.text not in valid:
        await message.answer("Выберите статью из списка:", reply_markup=articles_kb(valid))
        return
    article    = message.text
    await state.update_data(article=article, table_name=table_name)
    if data.get("from_pdf"):
        user_id = message.from_user.id
        session = pdf_sessions.get(user_id)
        if session:
            session["transactions"][data["pdf_idx"]]["article"]    = article
            session["transactions"][data["pdf_idx"]]["table_name"] = table_name
        updated_data = await state.get_data()
        await TransactionForm.final_confirmation.set()
        await message.answer(build_preview(updated_data), parse_mode="HTML", reply_markup=confirmation_kb())
        return
    if data.get("from_webhook"):
        updated_data = await state.get_data()
        await TransactionForm.final_confirmation.set()
        await message.answer(build_preview(updated_data), parse_mode="HTML", reply_markup=confirmation_kb())
        return
    await TransactionForm.amount.set()
    await message.answer("Введите сумму:", reply_markup=back_kb())

# ============================================================
# Сумма
# ============================================================
@dp.message_handler(state=TransactionForm.amount)
async def process_amount(message: types.Message, state: FSMContext):
    data = await state.get_data()
    if message.text == "⏪ Назад":
        tx_type    = data.get("tx_type", "Расход")
        table_name = data.get("table_name", "")
        articles   = INCOME_BY_TABLE.get(table_name, []) if tx_type == "Доход" else EXPENSE_BY_TABLE.get(table_name, [])
        await TransactionForm.article_choice.set()
        await message.answer("Выберите статью:", reply_markup=articles_kb(articles))
        return
    try:
        amount = float(message.text.replace(",", ".").replace(" ", ""))
        if amount <= 0:
            raise ValueError
    except ValueError:
        await message.answer("Введите корректную сумму (например: 5000 или 5000.50):", reply_markup=back_kb())
        return
    await state.update_data(amount=amount)
    await TransactionForm.currency.set()
    await message.answer("Выберите валюту:", reply_markup=currencies_kb())

# ============================================================
# Валюта
# ============================================================
@dp.message_handler(state=TransactionForm.currency)
async def process_currency(message: types.Message, state: FSMContext):
    if message.text == "⏪ Назад":
        await TransactionForm.amount.set()
        await message.answer("Введите сумму:", reply_markup=back_kb())
        return
    if message.text not in CURRENCIES:
        await message.answer("Выберите валюту из списка:", reply_markup=currencies_kb())
        return
    currency   = message.text
    data       = await state.get_data()
    rate       = get_cbr_rate(currency)
    amount_rub = round(float(data.get("amount", 0)) * rate, 2)
    await state.update_data(currency=currency, rate=rate, amount_rub=amount_rub)
    await TransactionForm.date.set()
    today = datetime.now().strftime("%d.%m.%Y, %H:%M")
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    kb.add(today)
    kb.add("⏪ Назад")
    await message.answer("Введите дату и время или нажмите кнопку:", reply_markup=kb)

# ============================================================
# Дата
# ============================================================
@dp.message_handler(state=TransactionForm.date)
async def process_date(message: types.Message, state: FSMContext):
    if message.text == "⏪ Назад":
        await TransactionForm.currency.set()
        await message.answer("Выберите валюту:", reply_markup=currencies_kb())
        return
    date_str = message.text.strip()
    parsed   = False
    for fmt in ("%d.%m.%Y, %H:%M", "%d.%m.%Y,%H:%M", "%d.%m.%Y"):
        try:
            datetime.strptime(date_str, fmt)
            parsed = True
            break
        except ValueError:
            continue
    if not parsed:
        await message.answer("Введите дату в формате ДД.ММ.ГГГГ, ЧЧ:ММ (например: 09.03.2026, 14:35):", reply_markup=back_kb())
        return
    await state.update_data(date=date_str)
    data = await state.get_data()
    await TransactionForm.final_confirmation.set()
    await message.answer(build_preview(data), parse_mode="HTML", reply_markup=confirmation_kb())

# ============================================================
# Подтверждение
# ============================================================
@dp.message_handler(state=TransactionForm.final_confirmation)
async def final_confirmation(message: types.Message, state: FSMContext):
    data = await state.get_data()
    if message.text == "✅ Записать":
        try:
            await save_transaction_to_sheets(data)
            await notify_admin(
                f"Таблица: {data.get('table_name', '')}\n"
                f"Статья: {data.get('article', '')}\n"
                f"Сумма: {float(data.get('amount', 0)):,.2f} {data.get('currency', 'RUB')}\n"
                f"В рублях: {float(data.get('amount_rub', data.get('amount', 0))):,.2f} ₽",
                message.from_user
            )
            if data.get("from_pdf"):
                user_id  = message.from_user.id
                session  = pdf_sessions.get(user_id)
                next_idx = data.get("pdf_idx", 0) + 1
                if session:
                    session["saved_count"] = session.get("saved_count", 0) + 1
                await state.finish()
                await PDFForm.reviewing.set()
                await message.answer("✅ Записано!", reply_markup=types.ReplyKeyboardRemove())
                await show_pdf_transaction(message, user_id, next_idx)
                return
            await message.answer("✅ Транзакция записана!", reply_markup=main_menu_kb())
            await state.finish()
            await TransactionForm.waiting_for_action.set()
        except Exception as e:
            logging.error(f"Ошибка записи: {e}")
            await message.answer(f"❌ Ошибка: {e}", reply_markup=main_menu_kb())
            await state.finish()
    elif message.text == "✏️ Изменить статью":
        await TransactionForm.table_choice.set()
        await message.answer("Выберите таблицу:", reply_markup=table_choice_kb(data.get("tx_type", "Расход")))
    elif message.text == "🔢 Изменить сумму/валюту":
        await TransactionForm.edit_amount.set()
        symbol = CURRENCY_SYMBOLS.get(data.get("currency","RUB"), data.get("currency","RUB"))
        await message.answer(
            f"Текущая сумма: <b>{data.get('amount',0)} {symbol}</b>\nВведите новую сумму:",
            parse_mode="HTML", reply_markup=back_kb()
        )
    elif message.text == "❌ Отменить":
        user_id = message.from_user.id
        if data.get("from_pdf"):
            idx = data.get("pdf_idx", 0)
            await state.finish()
            await PDFForm.reviewing.set()
            await message.answer("Отменено.", reply_markup=types.ReplyKeyboardRemove())
            await show_pdf_transaction(message, user_id, idx)
            return
        await state.finish()
        await TransactionForm.waiting_for_action.set()
        await message.answer("Операция отменена.", reply_markup=main_menu_kb())

# ============================================================
# Webhook — быстрые статьи (inline кнопки)
# ============================================================
@dp.callback_query_handler(lambda c: c.data.startswith("wbq|"), state="*")
async def process_webhook_quick(callback: types.CallbackQuery, state: FSMContext):
    parts      = callback.data.split("|")
    tx_id      = parts[1]
    article    = parts[2]
    table_name = parts[3] if len(parts) > 3 else ""
    tx = pending_transactions.pop(tx_id, None)
    if not tx:
        await callback.message.edit_text("❌ Транзакция устарела.")
        await callback.answer()
        return
    await state.update_data(
        amount=tx["a"], currency=tx["cur"], rate=tx["rate"],
        amount_rub=tx["a_rub"], date=tx["d"],
        comment=tx["m"], from_webhook=True, article=article,
        table_name=table_name,
        tx_type=tx.get("tx_type", "Расход"),
    )
    try:
        await callback.message.edit_reply_markup(reply_markup=None)
    except Exception:
        pass
    data = await state.get_data()
    await TransactionForm.final_confirmation.set()
    await callback.message.answer(build_preview(data), parse_mode="HTML", reply_markup=confirmation_kb())
    await callback.answer()

@dp.callback_query_handler(lambda c: c.data.startswith("wb|"), state="*")
async def process_webhook_callback(callback: types.CallbackQuery, state: FSMContext):
    payload = callback.data[3:]
    if payload == "no":
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except Exception:
            pass
        await callback.message.answer("❌ Пропущено.")
        await callback.answer()
        return
    tx = pending_transactions.pop(payload, None)
    if not tx:
        await callback.message.edit_text("❌ Транзакция устарела.")
        await callback.answer()
        return
    await state.update_data(
        amount=tx.get("a", 0),
        currency=tx.get("cur", "RUB"),
        rate=tx.get("rate", 1.0),
        amount_rub=tx.get("a_rub", tx.get("a", 0)),
        date=tx.get("d", datetime.now().strftime("%d.%m.%Y, %H:%M")),
        comment=tx.get("m", ""),
        tx_type=tx.get("tx_type", "Расход"),
        from_webhook=True,
    )
    try:
        await callback.message.edit_reply_markup(reply_markup=None)
    except Exception:
        pass
    tx_type = tx.get("tx_type", "Расход")
    await callback.message.answer(
        f"✅ Продолжаем: {tx.get('a')} {CURRENCY_SYMBOLS.get(tx.get('cur','RUB'),'₽')} — {tx.get('m', '')}\nВыберите таблицу:",
        reply_markup=table_choice_kb(tx_type)
    )
    await TransactionForm.table_choice.set()
    await callback.answer()

# ============================================================
# Мои транзакции
# ============================================================
@dp.message_handler(lambda m: m.text == "📋 Мои транзакции", state="*")
async def my_transactions(message: types.Message):
    try:
        ws      = sh.worksheet("Транзакции")
        records = ws.get_all_records()
        if not records:
            await message.answer("📂 Нет транзакций.", reply_markup=main_menu_kb())
            return
        last_10 = records[-10:]
        text = "📋 <b>Последние 10 транзакций:</b>\n"
        for rec in last_10:
            currency = str(rec.get("Валюта", "RUB")).strip() or "RUB"
            symbol   = CURRENCY_SYMBOLS.get(currency, currency)
            try:
                amount = float(str(rec.get("Сумма", 0)).replace(",", ".").replace(" ", "") or 0)
            except Exception:
                amount = 0
            try:
                raw_rub    = rec.get("Сумма в Руб", rec.get("Сумма в RUB", ""))
                amount_rub = float(str(raw_rub).replace(",", ".").replace(" ", "") or amount)
            except Exception:
                amount_rub = amount
            icon = "💰" if rec.get("Тип") == "Доход" else "💸"
            text += f"{icon} <b>{rec.get('Статья','')}</b>\n"
            text += f"  📅 {rec.get('Дата','')} | 📂 {rec.get('Таблица','')}\n"
            text += f"  💰 {amount:,.0f} {symbol}"
            if currency != "RUB":
                text += f" ({amount_rub:,.0f} ₽)"
            text += f"\n{'─' * 32}\n"
        await message.answer(text, parse_mode="HTML", reply_markup=main_menu_kb())
    except Exception as e:
        logging.error(f"Ошибка чтения транзакций: {e}")
        await message.answer("❌ Ошибка при загрузке транзакций.", reply_markup=main_menu_kb())

# ============================================================
# Статистика
# ============================================================
@dp.message_handler(lambda m: m.text == "📊 Статистика", state="*")
async def statistics(message: types.Message):
    try:
        ws      = sh.worksheet("Транзакции")
        records = ws.get_all_records()
        if not records:
            await message.answer("📂 Нет данных для статистики.", reply_markup=main_menu_kb())
            return
        current_month = datetime.now().strftime("%m.%Y")
        table_totals  = {}
        total_in = total_out = 0.0
        for rec in records:
            date_part = str(rec.get("Дата", "")).split(",")[0].strip()
            try:
                dt = datetime.strptime(date_part, "%d.%m.%Y")
                if dt.strftime("%m.%Y") == current_month:
                    tbl = rec.get("Таблица", "Прочее")
                    raw = rec.get("Сумма в Руб", rec.get("Сумма в RUB", rec.get("Сумма", 0)))
                    amt = float(str(raw).replace(",", ".").replace(" ", "") or 0)
                    table_totals[tbl] = table_totals.get(tbl, 0) + amt
                    if rec.get("Тип") == "Доход":
                        total_in += amt
                    else:
                        total_out += amt
            except (ValueError, TypeError):
                continue
        if not table_totals:
            await message.answer("📂 Нет транзакций за текущий месяц.", reply_markup=main_menu_kb())
            return
        total = sum(table_totals.values())
        icons = {
            "Поступления":     "📥",
            "Движение активов":      "🏦",
            "Крупные расходы": "💳",
            "Расходы":         "🛒",
        }
        text = f"📊 <b>Статистика за {current_month}:</b>\n"
        for tbl, amt in sorted(table_totals.items(), key=lambda x: x[1], reverse=True):
            pct = (amt / total) * 100 if total else 0
            bar = "█" * int(pct / 5) + "░" * (20 - int(pct / 5))
            text += f"{icons.get(tbl,'📂')} <b>{tbl}</b>\n{bar} {pct:.1f}%\n{amt:,.0f} ₽\n"
        text += (
            f"{'═' * 32}\n"
            f"📈 Доходы: {total_in:,.0f} ₽\n"
            f"📉 Расходы: {total_out:,.0f} ₽\n"
            f"💰 <b>ИТОГО:</b> {total:,.0f} ₽"
        )
        await message.answer(text, parse_mode="HTML", reply_markup=main_menu_kb())
    except Exception as e:
        logging.error(f"Ошибка статистики: {e}")
        await message.answer("❌ Ошибка при расчёте статистики.", reply_markup=main_menu_kb())

# ============================================================
# Настройки
# ============================================================
@dp.message_handler(lambda m: m.text == "⚙️ Настройки", state="*")
async def settings(message: types.Message):
    await message.answer(
        f"⚙️ <b>Настройки DizelFinance</b>\n"
        f"👤 Ваш ID: <code>{message.from_user.id}</code>\n"
        f"🗄 Google Sheets: {'✅' if sh else '❌'}\n"
        f"🤖 Gemini: {'✅ ' + GEMINI_MODEL if GEMINI_API_KEY else '❌ Не настроен'}\n"
        f"🌐 Proxy: {CLOUDFLARE_PROXY}\n"
        f"📅 Текущий лист: <b>{MONTH_SHEETS[datetime.now().month]}</b>\n"
        f"<b>Таблицы месяца:</b>\n"
        f"📥 Поступления      → col B / L\n"
        f"💳 Крупные расходы  → col O / Y\n"
        f"🛒 Расходы          → col AB / AL\n"
        f"🏦 Движение активов       → col BB / BL",
        parse_mode="HTML", reply_markup=main_menu_kb()
    )

# ============================================================
# Отложенные транзакции
# ============================================================
@dp.message_handler(lambda m: m.text == "📥 Отложенные", state="*")
async def show_drafts(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    drafts  = drafts_get(user_id)
    if not drafts:
        await message.answer("📭 Нет отложенных транзакций.", reply_markup=main_menu_kb())
        return
    text = f"📥 <b>Отложенные транзакции ({len(drafts)}):</b>\n"
    kb   = types.InlineKeyboardMarkup(row_width=1)
    for i, d in enumerate(drafts):
        symbol = CURRENCY_SYMBOLS.get(d["cur"], d["cur"])
        text  += f"{i + 1}. 💰 {d['a']:,.0f} {symbol} — {d['m']} ({d['d']})\n"
        kb.add(types.InlineKeyboardButton(
            f"✏️ #{i + 1} {d['m']} {d['a']:,.0f} {symbol}",
            callback_data=f"draft|{d['id']}"
        ))
    kb.add(types.InlineKeyboardButton("🗑 Очистить все", callback_data="draft|clear"))
    await message.answer(text, parse_mode="HTML", reply_markup=kb)

@dp.callback_query_handler(lambda c: c.data.startswith("draft|"), state="*")
async def process_draft(callback: types.CallbackQuery, state: FSMContext):
    user_id  = callback.from_user.id
    action   = callback.data.split("|")[1]
    if action == "clear":
        drafts_clear(user_id)
        await callback.message.edit_text("🗑 Все черновики удалены.")
        await callback.answer()
        return
    draft_id = action
    drafts   = drafts_get(user_id)
    draft    = next((d for d in drafts if d["id"] == draft_id), None)
    if not draft:
        await callback.message.edit_text("❌ Черновик не найден.")
        await callback.answer()
        return
    drafts_remove(draft_id)
    tx_type = draft.get("tx_type", "Расход")
    await state.update_data(
        amount=draft["a"],
        currency=draft["cur"],
        rate=draft["rate"],
        amount_rub=draft["a_rub"],
        date=draft["d"],
        comment=draft["m"],
        tx_type=tx_type,
        from_webhook=True,
    )
    symbol = CURRENCY_SYMBOLS.get(draft["cur"], "₽")
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.message.answer(
        f"✅ Продолжаем: {draft['a']:,.0f} {symbol} — {draft['m']}\nВыберите таблицу:",
        reply_markup=table_choice_kb(tx_type)
    )
    await TransactionForm.table_choice.set()
    await callback.answer()

# ============================================================
# Обработка фото (скриншот банка)
# ============================================================
@dp.message_handler(content_types=types.ContentType.PHOTO, state="*")
async def handle_screenshot(message: types.Message, state: FSMContext):
    if ALLOWED_IDS and message.from_user.id not in ALLOWED_IDS:
        return
    await message.answer("📸 Анализирую скриншот...")
    try:
        photo       = message.photo[-1]
        file        = await bot.get_file(photo.file_id)
        file_bytes  = await bot.download_file(file.file_path)
        image_bytes = file_bytes.read()
        transactions = parse_screenshot_transactions(image_bytes)
        if not transactions:
            await message.answer(
                "❌ Не удалось найти транзакции на скриншоте.\n"
                "Попробуйте:\n— Скриншот списка операций (не главного экрана)\n"
                "— Более чёткое фото\n— PDF выписку",
                reply_markup=main_menu_kb()
            )
            return
        if len(transactions) == 1:
            await _send_single_tx(message, transactions[0])
        else:
            existing = get_existing_transactions()
            await message.answer(f"📊 Найдено {len(transactions)} транзакций. Определяю статьи...")
            article_results = guess_articles_batch(transactions)
            enriched = _enrich_transactions(transactions, article_results, existing)
            _store_session(message.from_user.id, enriched)
            await _send_session_summary(message, enriched, "Скриншот")
    except Exception as e:
        logging.error(f"Ошибка обработки скриншота: {e}")
        await message.answer(f"❌ Ошибка: {e}", reply_markup=main_menu_kb())

async def _send_single_tx(message: types.Message, tx: dict):
    amount     = float(tx.get("amount", 0))
    currency   = tx.get("currency", "RUB")
    merchant   = tx.get("merchant", "")
    tx_type_w  = tx.get("tx_type", "Расход")
    date_raw   = tx.get("date", "")
    date       = date_raw if date_raw else datetime.now().strftime("%d.%m.%Y, %H:%M")
    hint       = tx.get("category_hint", "")
    rate       = get_cbr_rate(currency)
    amount_rub = round(amount * rate, 2)
    symbol     = CURRENCY_SYMBOLS.get(currency, currency)
    tx_icon    = "💰" if tx_type_w == "Доход" else "💸"
    article, table_name = guess_article(merchant, amount, tx_type=tx_type_w, hint=hint)
    tx_id = str(uuid.uuid4())[:8]
    pending_transactions[tx_id] = {
        "a": amount, "m": merchant, "d": date,
        "cur": currency, "rate": rate, "a_rub": amount_rub,
        "tx_type": tx_type_w, "article": article, "table_name": table_name,
    }
    drafts_add(message.from_user.id, {
        "id": str(uuid.uuid4())[:8], "a": amount, "m": merchant, "d": date,
        "cur": currency, "rate": rate, "a_rub": amount_rub, "tx_type": tx_type_w,
    })
    kb = types.InlineKeyboardMarkup(row_width=1)
    kb.add(types.InlineKeyboardButton(
        f"🤖 {table_name} → {article}", callback_data=f"wbq|{tx_id}|{article}|{table_name}"
    ))
    if tx_type_w == "Расход":
        alts = [a for a in EXPENSE_BY_TABLE.get(table_name, []) if a != article][:3]
    else:
        alts = [a for a in INCOME_BY_TABLE.get(table_name, []) if a != article][:3]
    for alt in alts:
        kb.add(types.InlineKeyboardButton(alt, callback_data=f"wbq|{tx_id}|{alt}|{table_name}"))
    kb.add(
        types.InlineKeyboardButton("📋 Все статьи", callback_data=f"wb|{tx_id}"),
        types.InlineKeyboardButton("❌ Пропустить",  callback_data="wb|no")
    )
    hint_line = f"\n💡 Банк: <i>{hint}</i>" if hint else ""
    preview = (
        f"📸 <b>Распознано:</b>\n"
        f"{tx_icon} {tx_type_w}\n"
        f"💵 {amount:,.2f} {symbol}\n"
        f"🏪 {merchant}\n"
        f"📅 {date}{hint_line}\n"
        f"📂 <b>{table_name}</b> → <b>{article}</b>\n"
        f"Подтвердите или выберите другую:"
    )
    await message.answer(preview, parse_mode="HTML", reply_markup=kb)

# ============================================================
# Обработка SMS текстом в боте
# ============================================================
@dp.message_handler(
    lambda m: m.text and len(m.text) > 20 and any(
        w in m.text.lower() for w in
        ["списано", "зачислено", "покупка", "оплата", "перевод", "баланс", "карта", "тенге", "рублей"]
    ),
    state=TransactionForm.waiting_for_action
)
async def handle_sms_text(message: types.Message, state: FSMContext):
    if ALLOWED_IDS and message.from_user.id not in ALLOWED_IDS:
        return
    await message.answer("📱 Похоже на банковское SMS, разбираю...")
    tx = parse_sms_transaction(message.text)
    if not tx:
        await message.answer("❌ Не смог распознать транзакцию. Попробуйте ➕ Новая транзакция.")
        return
    await _send_single_tx(message, tx)

# ============================================================
# Flask webhook
# ============================================================
app = Flask(__name__)

def _send_tg(chat_id, text, kb=None):
    payload = {"chat_id": chat_id, "text": text, "parse_mode": "HTML"}
    if kb:
        payload["reply_markup"] = kb.to_python()
    requests.post(f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage", json=payload)

def _build_quick_kb(tx_id, article, table_name, tx_type_w):
    kb = types.InlineKeyboardMarkup(row_width=1)
    kb.add(types.InlineKeyboardButton(
        f"🤖 {table_name} → {article}", callback_data=f"wbq|{tx_id}|{article}|{table_name}"
    ))
    if tx_type_w == "Расход":
        alts = [a for a in EXPENSE_BY_TABLE.get(table_name, []) if a != article][:3]
    else:
        alts = [a for a in INCOME_BY_TABLE.get(table_name, []) if a != article][:3]
    for alt in alts:
        kb.add(types.InlineKeyboardButton(alt, callback_data=f"wbq|{tx_id}|{alt}|{table_name}"))
    kb.add(
        types.InlineKeyboardButton("📋 Все статьи", callback_data=f"wb|{tx_id}"),
        types.InlineKeyboardButton("❌ Пропустить",  callback_data="wb|no")
    )
    return kb

@app.route("/webhook/transaction", methods=["POST"])
def webhook_transaction():
    try:
        data      = request.json
        user_id   = data.get("user_id")
        if not user_id or (ALLOWED_IDS and int(user_id) not in ALLOWED_IDS):
            return jsonify({"status": "error", "message": "Unauthorized"}), 403
        amount    = float(data.get("amount", 0))
        currency  = data.get("currency", "RUB")
        merchant  = data.get("merchant", "Неизвестно")
        date      = data.get("date", datetime.now().strftime("%d.%m.%Y, %H:%M"))
        tx_type_w = data.get("tx_type", "Расход")
        rate      = get_cbr_rate(currency)
        amount_rub = round(amount * rate, 2)
        symbol    = CURRENCY_SYMBOLS.get(currency, currency)
        icon      = "💰" if tx_type_w == "Доход" else "💸"
        article, table_name = guess_article(merchant, amount, tx_type=tx_type_w)
        tx_id   = str(uuid.uuid4())[:8]
        uid_int = int(user_id)
        pending_transactions[tx_id] = {
            "a": amount, "m": merchant, "d": date,
            "cur": currency, "rate": rate, "a_rub": amount_rub,
            "tx_type": tx_type_w, "article": article, "table_name": table_name,
        }
        drafts_add(uid_int, {
            "id": str(uuid.uuid4())[:8], "a": amount, "m": merchant, "d": date,
            "cur": currency, "rate": rate, "a_rub": amount_rub, "tx_type": tx_type_w,
        })
        text = (
            f"🔔 <b>Новая транзакция:</b>\n{icon} {tx_type_w}\n"
            f"💰 {amount:,.2f} {symbol}"
            + (f"\n🔄 {amount_rub:,.2f} ₽" if currency != "RUB" else "")
            + f"\n🏪 {merchant}\n📅 {date}\n"
            f"🤖 {table_name} → {article}\nЗаписать?"
        )
        kb = _build_quick_kb(tx_id, article, table_name, tx_type_w)
        _send_tg(user_id, text, kb)
        return jsonify({"status": "ok"}), 200
    except Exception as e:
        logging.error(f"Ошибка webhook: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/webhook/sms", methods=["POST"])
def webhook_sms():
    try:
        data     = request.json
        logging.info(f"SMS WEBHOOK: {data}")
        user_id  = data.get("user_id")
        sms_text = data.get("sms", "").strip()
        if not user_id or not sms_text:
            return jsonify({"status": "error", "message": "Missing user_id or sms"}), 400
        if ALLOWED_IDS and int(user_id) not in ALLOWED_IDS:
            return jsonify({"status": "error", "message": "Unauthorized"}), 403
        tx = parse_sms_transaction(sms_text)
        if not tx:
            return jsonify({"status": "skip", "message": "Not a transaction SMS"}), 200
        amount    = float(tx.get("amount", 0))
        currency  = tx.get("currency", "RUB")
        merchant  = tx.get("merchant", "SMS")
        tx_type_w = tx.get("tx_type", "Расход")
        date      = tx.get("date") or datetime.now().strftime("%d.%m.%Y, %H:%M")
        rate      = get_cbr_rate(currency)
        amount_rub = round(amount * rate, 2)
        symbol    = CURRENCY_SYMBOLS.get(currency, currency)
        icon      = "💰" if tx_type_w == "Доход" else "💸"
        article, table_name = guess_article(merchant, amount, tx_type=tx_type_w)
        tx_id   = str(uuid.uuid4())[:8]
        uid_int = int(user_id)
        pending_transactions[tx_id] = {
            "a": amount, "m": merchant, "d": date,
            "cur": currency, "rate": rate, "a_rub": amount_rub,
            "tx_type": tx_type_w, "article": article, "table_name": table_name,
        }
        drafts_add(uid_int, {
            "id": str(uuid.uuid4())[:8], "a": amount, "m": merchant, "d": date,
            "cur": currency, "rate": rate, "a_rub": amount_rub, "tx_type": tx_type_w,
        })
        text = (
            f"📱 <b>SMS транзакция:</b>\n{icon} {tx_type_w}\n"
            f"💵 {amount:,.2f} {symbol}"
            + (f"\n🔄 {amount_rub:,.2f} ₽" if currency != "RUB" else "")
            + f"\n🏪 {merchant}\n📅 {date}\n"
            f"🤖 {table_name} → {article}\nЗаписать?"
        )
        kb = _build_quick_kb(tx_id, article, table_name, tx_type_w)
        _send_tg(user_id, text, kb)
        return jsonify({"status": "ok"}), 200
    except Exception as e:
        logging.error(f"Ошибка SMS webhook: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "model": GEMINI_MODEL, "sheet": "connected" if sh else "error"}), 200

# ============================================================
# Неизвестные сообщения
# ============================================================
@dp.message_handler(state="*")
async def unknown_message(message: types.Message, state: FSMContext):
    current = await state.get_state()
    if current is None or current == TransactionForm.waiting_for_action.state:
        await message.answer("Не понимаю. Используйте меню 👇", reply_markup=main_menu_kb())

# ============================================================
# Запуск
# ============================================================
def run_bot():
    executor.start_polling(dp, skip_updates=True)

def run_flask():
    app.run(host="0.0.0.0", port=5000, use_reloader=False)

if __name__ == "__main__":
    flask_thread = threading.Thread(target=run_flask, daemon=True)
    flask_thread.start()
    logging.info("🚀 DizelFinance Bot v2 запущен! (Claude API + Planergo Structure)")
    run_bot()